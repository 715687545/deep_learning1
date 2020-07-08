from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import xlrd
import xlwt
try:
    filepath=input("请输入excel文件路径，请注意文件后缀为.XLSX\n")	
    #wb=load_workbook(filename = r'D:/深圳水务工程检测/检测数据处理/6-23/检测数据与截图/福田区排水管网正本清源工程（第六期）勘察设计施工项目总承包（EPC）第六批次6.18/上围村/抽检内窥检测记录表-上围村.XLSX')		#打开目标excel
    wb=load_workbook(filename = filepath)


    #print(wb)
    sh=wb["Sheet1"]				#选取表单1
    maxcolumn=sh.max_row			#获取最大列总数
    print(maxcolumn)
    #etst=input()
    styleBlueBkg = xlwt.easyxf('pattern: pattern solid, fore_colour red;')  # 给单元格上底色红色
    orange_fill = PatternFill(fill_type='solid', fgColor="FFC125")

    count=0

    for i in range(1,maxcolumn):
    
        ce=sh.cell(row=i,column=2)		###读取第二行第i列数据
        de=sh.cell(row=i,column=3)		#读取第三行第i列数据
        if not de.value==None:
            for j in range(i,maxcolumn):
                fe=sh.cell(row=j,column=3)	#读取第三列第j行数据
                if ce.value==fe.value:			#如果起止点编号相同，检查是否为反向
                    ge=sh.cell(row=j,column=2)#读取第二列第j行数据
                    if de.value==ge.value:
                        count+=1
                        print("第",i,"管段为反向视频")
                        sh.cell(row=i,column=2).fill = orange_fill
                        sh.cell(row=j,column=2).fill = orange_fill
except:
    print("发生未知错误")
wb.save(filepath)
wb.close
print("已标注反向管段，共有反向视频",count,"段\n")
input("请按enter键结束程序")
